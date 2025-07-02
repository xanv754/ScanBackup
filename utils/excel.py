import pandas as pd
from typing import Dict
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import Font, PatternFill, Border, Side, numbers
from constants import cells, HeaderDailyReport
from utils.translate import Translate

class ExcelExport:
    """Class to export data to excel."""

    filepath: str
    data: Dict[str, pd.DataFrame]

    def __init__(self, filename: str, data: Dict[str, pd.DataFrame]) -> None:
        self.__set_filepath(filename=filename)
        self.data = data


    def __set_filepath(self, filename: str) -> None:
        """Set the filepath to export data."""
        if not filename.endswith(".xlsx"): filename = f"{filename}.xlsx"
        home = Path.home()
        if Path(home / "Downloads").exists():
            self.filepath = f"{home / 'Downloads'}/{filename}"
        elif Path(home / "Descargas").exists():
            self.filepath = f"{home / 'Descargas'}/{filename}"
        else:
            self.filepath = f"{home}/{filename}"

    def __set_styles(self) -> None:
        """Set the styles to excel."""
        border = Border(left=Side(style="thin", color=Color(rgb="000000")), right=Side(style="thin", color=Color(rgb="000000")), top=Side(style="thin", color=Color(rgb="000000")), bottom=Side(style="thin", color=Color(rgb="000000")))
        number_format = '0.00'

        workbook = load_workbook(self.filepath)
        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
            max_column = sheet.max_column
            max_row = sheet.max_row

            sheet.column_dimensions[cells[1]].width = 57
            for column in range(2, max_column + 1):
                sheet.column_dimensions[cells[column]].width = 16
                sheet.freeze_panes = cells[column] + str(2)

            bg = PatternFill(fill_type="solid", start_color=Color(rgb="16365C"), end_color=Color(rgb="16365C"))
            font = Font(name="Liberation Sans", size=11, bold=True, color=Color(rgb="FFFFFF"))
            for column in range(1, max_column + 1):
                sheet.cell(row=1, column=column).font = font
                sheet.cell(row=1, column=column).fill = bg
                sheet.cell(row=1, column=column).border = border

            font = Font(name="Liberation Sans", bold=False, color=Color(rgb="000000"))
            for row in range(2, max_row + 1):
                for column in range(1, max_column + 1):
                    sheet.cell(row=row, column=column).font = font
                    sheet.cell(row=row, column=column).border = border
                    if (cells[column] == "E" or
                        cells[column] == "F" or
                        cells[column] == "G" or
                        cells[column] == "H" or
                        cells[column] == "I"
                    ): sheet.cell(row=row, column=column).number_format = number_format

        workbook.save(self.filepath)

    def export(self) -> None:
        """Export data to excel."""
        with pd.ExcelWriter(self.filepath, engine="openpyxl") as writer:
            for layer_type, df in self.data.items():
                df.sort_values(by=[HeaderDailyReport.TYPE, HeaderDailyReport.NAME], inplace=True)
                df = Translate.header(df)
                df.to_excel(writer, sheet_name=layer_type, index=False)
        self.__set_styles()


if __name__ == "__main__":
    data = {
        "BORDE": pd.DataFrame({
            "Fecha": ["2025-01-01", "2025-01-02", "2025-01-03"],
            "Prom": ["10", "20", "30"],
            "MaxProm": ["10", "20", "30"]
        }),
    }
    excel = ExcelExport(filename="test", data=data)
    excel.export()
    print("Exported in", excel.filepath)
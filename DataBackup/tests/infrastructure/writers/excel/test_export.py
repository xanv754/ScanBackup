from openpyxl import load_workbook
from pydantic import BaseModel
from scanbackup.infrastructure.writers.excel.export import ExcelWriter
from scanbackup.shared import ExcelExportError
from tests.support import TempDirTestCase


class _Row(BaseModel):
    """Minimal pydantic model used to exercise ExcelWriter.export."""

    layer: str
    name: str
    value: int


class TestExcelWriter(TempDirTestCase):
    """Unit tests for the ExcelWriter.export method."""

    def test_writes_one_sheet_per_name_and_returns_filepath(self) -> None:
        """export() must group rows by sheet_field and create one sheet per sheet_names entry."""
        writer = ExcelWriter(dir=self.tmp_dir)
        data = [
            _Row(layer="BORDE", name="a", value=1),
            _Row(layer="DINT", name="b", value=2),
        ]

        result = writer.export(
            filename="report",
            data=data,
            model=_Row,
            sheet_field="layer",
            sheet_names=["BORDE", "DINT"],
        )

        expected_path = self.tmp_dir / "report.xlsx"
        self.assertEqual(result, str(expected_path.resolve()))
        workbook = load_workbook(expected_path)
        self.assertEqual(workbook.sheetnames, ["BORDE", "DINT"])

    def test_creates_empty_sheet_with_headers_for_names_with_no_rows(self) -> None:
        """A sheet name with no matching rows must still be created, with the model's headers."""
        writer = ExcelWriter(dir=self.tmp_dir)
        data = [_Row(layer="BORDE", name="a", value=1)]

        writer.export(
            filename="report",
            data=data,
            model=_Row,
            sheet_field="layer",
            sheet_names=["BORDE", "DINT"],
        )

        workbook = load_workbook(self.tmp_dir / "report.xlsx")
        empty_sheet = workbook["DINT"]
        self.assertEqual(
            [cell.value for cell in empty_sheet[1]], ["layer", "name", "value"]
        )
        self.assertEqual(empty_sheet.max_row, 1)

    def test_excludes_given_fields_from_headers_and_rows(self) -> None:
        """export() must omit excluded fields from both headers and row values."""
        writer = ExcelWriter(dir=self.tmp_dir)
        data = [_Row(layer="BORDE", name="a", value=1)]

        writer.export(
            filename="report",
            data=data,
            model=_Row,
            sheet_field="layer",
            sheet_names=["BORDE"],
            exclude={"layer"},
        )

        workbook = load_workbook(self.tmp_dir / "report.xlsx")
        sheet = workbook["BORDE"]
        self.assertEqual([cell.value for cell in sheet[1]], ["name", "value"])

    def test_invalid_sheet_name_raises_excel_export_error(self) -> None:
        """A failure while writing the workbook must be wrapped into ExcelExportError."""
        writer = ExcelWriter(dir=self.tmp_dir)
        data = [_Row(layer="BORDE", name="a", value=1)]

        with self.assertRaises(ExcelExportError):
            writer.export(
                filename="report",
                data=data,
                model=_Row,
                sheet_field="layer",
                sheet_names=[],
            )


if __name__ == "__main__":
    import unittest

    unittest.main()

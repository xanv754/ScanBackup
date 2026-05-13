from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import Font, PatternFill, Border, Side
from scanbackup.domain import BaseExport
from scanbackup.shared import ExcelExportError

cells = {
    1: "A",
    2: "B",
    3: "C",
    4: "D",
    5: "E",
    6: "F",
    7: "G",
    8: "H",
    9: "I",
    10: "J",
    11: "K",
    12: "L",
    13: "M",
    14: "N",
    15: "O",
    16: "P",
    17: "Q",
    18: "R",
    19: "S",
    20: "T",
    21: "U",
    22: "V",
    23: "W",
    24: "X",
    25: "Y",
    26: "Z",
}


class ExcelExport(BaseExport):
    def _styles(self) -> None:
        """Set the styles to excel."""
        border = Border(
            left=Side(style="thin", color=Color(rgb="000000")),
            right=Side(style="thin", color=Color(rgb="000000")),
            top=Side(style="thin", color=Color(rgb="000000")),
            bottom=Side(style="thin", color=Color(rgb="000000")),
        )
        number_format = "0.00"

        workbook = load_workbook(self.filepath)
        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
            max_column = sheet.max_column
            max_row = sheet.max_row

            sheet.column_dimensions[cells[1]].width = 57
            for column in range(2, max_column + 1):
                sheet.column_dimensions[cells[column]].width = 16
                sheet.freeze_panes = cells[column] + str(2)

            bg = PatternFill(
                fill_type="solid",
                start_color=Color(rgb="16365C"),
                end_color=Color(rgb="16365C"),
            )
            font = Font(
                name="Liberation Sans", size=11, bold=True, color=Color(rgb="FFFFFF")
            )
            for column in range(1, max_column + 1):
                sheet.cell(row=1, column=column).font = font
                sheet.cell(row=1, column=column).fill = bg
                sheet.cell(row=1, column=column).border = border

            font = Font(name="Liberation Sans", bold=False, color=Color(rgb="000000"))
            for row in range(2, max_row + 1):
                for column in range(1, max_column + 1):
                    sheet.cell(row=row, column=column).font = font
                    sheet.cell(row=row, column=column).border = border
                    if cells[column] == "H":
                        sheet.cell(row=row, column=column).number_format = number_format
                    if (
                        cells[column] == "B"
                        or cells[column] == "C"
                        or cells[column] == "D"
                        or cells[column] == "E"
                        or cells[column] == "F"
                        or cells[column] == "G"
                        or cells[column] == "H"
                        or cells[column] == "I"
                    ):
                        sheet.cell(row=row, column=column).number_format = number_format

        workbook.save(self.filepath)

    def execute(self) -> bool:
        filepath = self.get_filepath()
        data = self.get_data()
        try:
            with ExcelWriter(filepath, engine="openpyxl") as writer:
                for layer_type, df in data.items():
                    df.to_excel(writer, sheet_name=layer_type, index=False)  # type: ignore
            self._styles()
        except Exception as error:
            ExcelExportError(error=error, filename=filepath)
            return False
        else:
            return True
